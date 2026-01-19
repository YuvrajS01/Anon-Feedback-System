"""
Flask-Based Local Anonymous Student Feedback System.

This application provides:
- Token-based anonymous feedback submission
- Multi-teacher feedback per token (admin configurable combos)
- Admin dashboard with analytics
- Excel report exports

Run with: python app.py
Access at: http://<local-ip>:5000
"""

import os
import logging
from io import BytesIO
from functools import wraps
from datetime import datetime

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, send_file, abort, jsonify
)
from flask_wtf.csrf import CSRFProtect
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

import config
from database import (
    init_db, validate_token, mark_token_used, save_feedback,
    get_token_stats, get_all_feedback, get_feedback_by_teacher,
    get_feedback_by_subject, get_teacher_summary, get_question_averages,
    create_session, get_session_by_token, update_session_progress,
    get_completed_combo_indices, get_session_stats
)


# Disable Flask access logs for anonymity
log = logging.getLogger('werkzeug')
log.setLevel(logging.ERROR)

# Initialize Flask app
app = Flask(__name__)
app.secret_key = config.SECRET_KEY
app.config['WTF_CSRF_ENABLED'] = True

# Enable CSRF protection
csrf = CSRFProtect(app)


def admin_required(f):
    """Decorator to require admin authentication."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function


def get_current_combos():
    """Get current teacher-subject combos (reload from config for live updates)."""
    return config.load_combos()


# =============================================================================
# Student Routes
# =============================================================================

@app.route('/')
def index():
    """Token entry page."""
    return render_template('index.html')


@app.route('/verify-token', methods=['POST'])
def verify_token():
    """Validate the submitted token and create/resume session."""
    token = request.form.get('token', '').strip().upper()
    
    if not token:
        flash('Please enter a token.', 'error')
        return redirect(url_for('index'))
    
    combos = get_current_combos()
    
    if not combos:
        flash('No feedback combos configured. Please contact admin.', 'error')
        return redirect(url_for('index'))
    
    # Check if session already exists for this token
    existing_session = get_session_by_token(token)
    
    if existing_session:
        if existing_session['is_complete']:
            flash('This token has already been used to submit all feedback.', 'error')
            return redirect(url_for('index'))
        # Resume existing session
        session['valid_token'] = token
        session['session_id'] = existing_session['id']
        session['total_combos'] = existing_session['total_combos']
        # Find next incomplete combo
        completed = get_completed_combo_indices(existing_session['id'])
        next_index = 0
        for i in range(len(combos)):
            if i not in completed:
                next_index = i
                break
        return redirect(url_for('feedback_step', index=next_index))
    
    # Validate token exists and is unused
    if not validate_token(token):
        flash('Invalid or already used token.', 'error')
        return redirect(url_for('index'))
    
    # Create new session
    session_id = create_session(token, len(combos))
    session['valid_token'] = token
    session['session_id'] = session_id
    session['total_combos'] = len(combos)
    
    return redirect(url_for('feedback_step', index=0))


@app.route('/feedback/<int:index>')
def feedback_step(index):
    """Feedback form page for a specific teacher-subject combo."""
    if 'valid_token' not in session or 'session_id' not in session:
        flash('Please enter a valid token first.', 'error')
        return redirect(url_for('index'))
    
    combos = get_current_combos()
    
    if index < 0 or index >= len(combos):
        flash('Invalid feedback step.', 'error')
        return redirect(url_for('index'))
    
    # Check if this combo is already completed
    completed = get_completed_combo_indices(session['session_id'])
    if index in completed:
        # Find next incomplete or go to thank you
        for i in range(len(combos)):
            if i not in completed:
                return redirect(url_for('feedback_step', index=i))
        return redirect(url_for('thankyou'))
    
    current_combo = combos[index]
    
    return render_template(
        'feedback_step.html',
        combo=current_combo,
        combo_index=index,
        total_combos=len(combos),
        completed_count=len(completed),
        questions=config.QUESTIONS
    )


@app.route('/submit/<int:index>', methods=['POST'])
def submit_step(index):
    """Process and save feedback for a specific combo."""
    if 'valid_token' not in session or 'session_id' not in session:
        flash('Session expired. Please enter your token again.', 'error')
        return redirect(url_for('index'))
    
    session_id = session['session_id']
    combos = get_current_combos()
    
    if index < 0 or index >= len(combos):
        flash('Invalid feedback step.', 'error')
        return redirect(url_for('index'))
    
    current_combo = combos[index]
    teacher = current_combo['teacher']
    subject = current_combo['subject']
    comment = request.form.get('comment', '').strip()
    
    # Collect ratings (q1 to q10)
    ratings = []
    for i in range(1, 11):
        try:
            rating = int(request.form.get(f'q{i}', 0))
            if rating < 1 or rating > 10:
                raise ValueError()
            ratings.append(rating)
        except (ValueError, TypeError):
            flash(f'Please provide a valid rating (1-10) for all questions.', 'error')
            return redirect(url_for('feedback_step', index=index))
    
    # Save feedback
    save_feedback(session_id, index, teacher, subject, ratings, comment)
    
    # Update session progress
    completed = get_completed_combo_indices(session_id)
    completed_count = len(completed)
    
    if completed_count >= len(combos):
        # All combos completed, mark token as used and session as complete
        mark_token_used(session['valid_token'])
        update_session_progress(session_id, completed_count, is_complete=True)
        # Clear session
        session.pop('valid_token', None)
        session.pop('session_id', None)
        session.pop('total_combos', None)
        return redirect(url_for('thankyou'))
    else:
        update_session_progress(session_id, completed_count)
        # Find next incomplete combo
        for i in range(len(combos)):
            if i not in completed:
                return redirect(url_for('feedback_step', index=i))
    
    return redirect(url_for('thankyou'))


@app.route('/thankyou')
def thankyou():
    """Thank you confirmation page."""
    return render_template('thankyou.html')


# =============================================================================
# Admin Routes
# =============================================================================

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    """Admin login page."""
    if request.method == 'POST':
        password = request.form.get('password', '')
        if password == config.ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Invalid password.', 'error')
    
    return render_template('admin_login.html')


@app.route('/admin/logout')
def admin_logout():
    """Admin logout."""
    session.pop('admin_logged_in', None)
    flash('Logged out successfully.', 'success')
    return redirect(url_for('admin_login'))


@app.route('/admin')
@admin_required
def admin_dashboard():
    """Admin dashboard with statistics and analytics."""
    token_stats = get_token_stats()
    session_stats = get_session_stats()
    teacher_summary = get_teacher_summary()
    question_averages = get_question_averages()
    combos = get_current_combos()
    
    # Get unique teachers and subjects from combos for export options
    teachers = list(set(c['teacher'] for c in combos))
    subjects = list(set(c['subject'] for c in combos))
    
    return render_template(
        'admin_dashboard.html',
        token_stats=token_stats,
        session_stats=session_stats,
        teacher_summary=teacher_summary,
        question_averages=question_averages,
        questions=config.QUESTIONS,
        teachers=teachers,
        subjects=subjects,
        combos=combos
    )


# =============================================================================
# Excel Export Routes
# =============================================================================

def create_excel_workbook(feedback_data: list, title: str) -> BytesIO:
    """Create an Excel workbook from feedback data."""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit
    
    # Headers
    headers = [
        'Teacher', 'Subject',
        'Q1', 'Q2', 'Q3', 'Q4', 'Q5', 'Q6', 'Q7', 'Q8', 'Q9', 'Q10',
        'Comment', 'Submitted At'
    ]
    
    # Write headers with bold font
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Write data
    for row_idx, entry in enumerate(feedback_data, 2):
        ws.cell(row=row_idx, column=1, value=entry['teacher'])
        ws.cell(row=row_idx, column=2, value=entry['subject'])
        for q in range(1, 11):
            ws.cell(row=row_idx, column=q + 2, value=entry[f'q{q}'])
        ws.cell(row=row_idx, column=13, value=entry.get('comment', ''))
        ws.cell(row=row_idx, column=14, value=entry.get('submitted_at', ''))
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = max(
            len(str(ws.cell(row=row, column=col).value or ''))
            for row in range(1, len(feedback_data) + 2)
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 50)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.route('/export/all')
@admin_required
def export_all():
    """Export all feedback as Excel."""
    feedback = get_all_feedback()
    if not feedback:
        flash('No feedback data to export.', 'error')
        return redirect(url_for('admin_dashboard'))
    
    excel_file = create_excel_workbook(feedback, 'All Feedback')
    filename = f'feedback_all_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/export/teacher/<teacher_name>')
@admin_required
def export_teacher(teacher_name: str):
    """Export feedback for a specific teacher."""
    feedback = get_feedback_by_teacher(teacher_name)
    if not feedback:
        flash(f'No feedback data for {teacher_name}.', 'error')
        return redirect(url_for('admin_dashboard'))
    
    excel_file = create_excel_workbook(feedback, teacher_name)
    safe_name = teacher_name.replace(' ', '_').replace('.', '')
    filename = f'feedback_{safe_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/export/subject/<subject_name>')
@admin_required
def export_subject(subject_name: str):
    """Export feedback for a specific subject."""
    feedback = get_feedback_by_subject(subject_name)
    if not feedback:
        flash(f'No feedback data for {subject_name}.', 'error')
        return redirect(url_for('admin_dashboard'))
    
    excel_file = create_excel_workbook(feedback, subject_name)
    safe_name = subject_name.replace(' ', '_')
    filename = f'feedback_{safe_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# =============================================================================
# Main Entry Point
# =============================================================================

if __name__ == '__main__':
    # Initialize database
    init_db()
    
    # Get local IP for display
    import socket
    hostname = socket.gethostname()
    try:
        local_ip = socket.gethostbyname(hostname)
    except:
        local_ip = '127.0.0.1'
    
    combos = get_current_combos()
    
    print("=" * 60)
    print("  Anonymous Student Feedback System")
    print("=" * 60)
    print(f"  Local URL:    http://127.0.0.1:5000")
    print(f"  Network URL:  http://{local_ip}:5000")
    print(f"  Admin Panel:  http://127.0.0.1:5000/admin")
    print(f"  Combos:       {len(combos)} teacher-subject pairs")
    print("=" * 60)
    print("  Press Ctrl+C to stop the server")
    print("=" * 60)
    
    # Run Flask app on all interfaces
    app.run(host='0.0.0.0', port=5000, debug=False)
